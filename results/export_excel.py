# Libraries
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Function
def export_to_excel(data:pd.DataFrame,data_quality:float = None, file_path: str = "data_quality_report.xlsx"):

    # Normalizar entrada
    if not isinstance(data, pd.DataFrame):
        df = pd.DataFrame(data, columns=["tabla", "campo", "dimension", "score", "fecha_ejecucion"])
    else:
        df = data.copy()

    df["score"] = df["score"].astype(float)

    # Estilos
    header_font   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    header_fill   = PatternFill("solid", start_color="2F5496")
    subtotal_fill = PatternFill("solid", start_color="D9E1F2")
    subtotal_font = Font(bold=True, name="Arial", size=10)
    cell_font     = Font(name="Arial", size=10)
    center        = Alignment(horizontal="center", vertical="center")
    thin          = Side(style="thin", color="AAAAAA")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(cell, value):
        cell.value     = value
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

    def style_data(cell, value, pct=False, bold=False):
        cell.value     = value
        cell.font      = subtotal_font if bold else cell_font
        cell.alignment = center
        cell.border    = border
        if bold:
            cell.fill = subtotal_fill
        if pct and isinstance(value, (int, float)):
            cell.number_format = '0.00"%"'

    def write_section_title(ws, row, text):
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, name="Arial", size=12)

    def autofit(ws):
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value is not None),
                default=8
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

    # Dimensiones en orden consistente
    dimension_order = ["completitud", "validez", "unicidad", "precision", "consistencia", "oportunidad"]
    existing_dims   = [d for d in dimension_order if d in df["dimension"].str.lower().unique()]
    rename_map      = {
        "completitud":  "Completitud",
        "validez":      "Validez",
        "unicidad":     "Unicidad",
        "precision":    "Precisión",
        "consistencia": "Consistencia",
        "oportunidad":  "Oportunidad",
    }
    dims_display = [rename_map[d] for d in existing_dims]

    df["dimension_lower"] = df["dimension"].str.lower()

    # Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Quality"

    # ──────────────────────────────────────────────
    # SCORE GLOBAL
    # ──────────────────────────────────────────────
    score_fill = PatternFill("solid", start_color="1F3864")  # azul oscuro

    UMBRAL_DAMA = 85

    df["punto_dama"] = (df["score"] > UMBRAL_DAMA).astype(int)

    puntos_obtenidos = int(df["punto_dama"].sum())
    puntos_posibles = int(len(df))
    score_dama = round((puntos_obtenidos / puntos_posibles) * 100, 2) if puntos_posibles > 0 else 0

    # Fila 1: Score oficial
    c_label = ws.cell(row=1, column=1, value="Calificación DAMA oficial")
    c_value = ws.cell(row=1, column=2, value=score_dama)

    c_label.font      = Font(bold=True, color="FFFFFF", name="Arial", size=13)
    c_label.fill      = score_fill
    c_label.alignment = center
    c_label.border    = border

    c_value.font          = Font(bold=True, color="FFFFFF", name="Arial", size=13)
    c_value.fill          = score_fill
    c_value.alignment     = center
    c_value.border        = border
    c_value.number_format = '0.00"%"'

    # Fila 2: Fórmula
    style_data(ws.cell(row=2, column=1), "Fórmula", bold=True)
    style_data(
        ws.cell(row=2, column=2),
        f"{puntos_obtenidos} puntos obtenidos / {puntos_posibles} puntos posibles"
    )

    # Fila 3: Umbral
    style_data(ws.cell(row=3, column=1), "Umbral de cumplimiento", bold=True)
    style_data(ws.cell(row=3, column=2), f"> {UMBRAL_DAMA}%")

    # Fila 5: Nota metodológica
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=len(existing_dims) + 2)

    nota = ws.cell(row=5, column=1)
    nota.value = (
        "Nota: La calificación oficial se calcula bajo el marco DAMA. "
        "Cada campo evaluado en cada dimensión obtiene 1 punto si supera el 85%. "
        "Los promedios mostrados en las tablas son referenciales y no representan la calificación oficial."
    )
    nota.font = Font(italic=True, name="Arial", size=10)
    nota.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    nota.border = border

    ws.row_dimensions[5].height = 35

    # ──────────────────────────────────────────────
    # TABLA 1: Detalle por Campo
    # ──────────────────────────────────────────────
    write_section_title(ws, 7, "Detalle por Campo")

    t1_headers = ["Campo", "Tabla"] + dims_display
    for col_idx, h in enumerate(t1_headers, start=1):
        style_header(ws.cell(row=8, column=col_idx), h)

    # Pivotear df para cruzar campo/tabla vs dimensión
    pivot_df = df.pivot_table(
        index=["campo", "tabla"],
        columns="dimension_lower",
        values="score",
        aggfunc="first"
    ).reset_index()

    for i, (_, row) in enumerate(pivot_df.iterrows()):
        excel_row = 9 + i
        style_data(ws.cell(excel_row, 1), row["campo"])
        style_data(ws.cell(excel_row, 2), row["tabla"])
        for col_idx, dim in enumerate(existing_dims, start=3):
            val = row.get(dim)
            style_data(
                ws.cell(excel_row, col_idx),
                round(val, 2) if pd.notna(val) else "-",
                pct=pd.notna(val)
            )
    # ──────────────────────────────────────────────
    # TABLA 2: Resumen por Tabla
    # ──────────────────────────────────────────────
    pivot_tabla = (
        df.groupby(["tabla", "dimension_lower"])["score"]
        .mean()
        .unstack(level="dimension_lower")
        .reindex(columns=existing_dims)
        .reset_index()
    )

    t2_start = ws.max_row + 3
    write_section_title(ws, t2_start - 1, "Resumen descriptivo por Tabla")

    for col_idx, h in enumerate(["Tabla"] + dims_display + ["Promedio aritmético referencial"], start=1):
        style_header(ws.cell(t2_start, col_idx), h)

    for i, (_, row) in enumerate(pivot_tabla.iterrows()):
        excel_row = t2_start + 1 + i
        style_data(ws.cell(excel_row, 1), row["tabla"])
        row_vals = []
        for col_idx, dim in enumerate(existing_dims, start=2):
            val = row.get(dim)
            display = round(val, 2) if pd.notna(val) else "-"
            style_data(ws.cell(excel_row, col_idx), display, pct=pd.notna(val))
            if pd.notna(val):
                row_vals.append(val)
        prom = round(sum(row_vals) / len(row_vals), 2) if row_vals else "-"
        style_data(ws.cell(excel_row, len(existing_dims) + 2), prom, pct=isinstance(prom, float), bold=True)

    # Footer promedios
    footer_t2 = t2_start + 1 + len(pivot_tabla)
    style_data(ws.cell(footer_t2, 1), "Promedio aritmético referencial", bold=True)
    all_avgs = []
    for col_idx, dim in enumerate(existing_dims, start=2):
        col_vals = pivot_tabla[dim].dropna().tolist()
        avg = round(sum(col_vals) / len(col_vals), 2) if col_vals else None
        style_data(ws.cell(footer_t2, col_idx), avg if avg is not None else "-", pct=avg is not None, bold=True)
        if avg is not None:
            all_avgs.append(avg)
    prom_gen = round(sum(all_avgs) / len(all_avgs), 2) if all_avgs else "-"
    style_data(ws.cell(footer_t2, len(existing_dims) + 2), prom_gen, pct=isinstance(prom_gen, float), bold=True)

    # ──────────────────────────────────────────────
    # TABLA 3: Resumen por Dimensión
    # ──────────────────────────────────────────────
    dim_avgs = (
        df.groupby("dimension_lower")["score"]
        .mean()
        .reindex(existing_dims)
    )

    t3_start = ws.max_row + 3
    write_section_title(ws, t3_start - 1, "Resumen descriptivo por Dimensión")

    for col_idx, h in enumerate(["Dimensión", "Promedio aritmético (%)"], start=1):
        style_header(ws.cell(t3_start, col_idx), h)

    for i, (dim, avg) in enumerate(dim_avgs.items(), start=1):
        style_data(ws.cell(t3_start + i, 1), rename_map.get(dim, dim))
        style_data(ws.cell(t3_start + i, 2), round(avg, 2), pct=True)

    footer_t3 = t3_start + len(dim_avgs) + 1
    style_data(ws.cell(footer_t3, 1), "Promedio aritmético referencial", bold=True)
    style_data(ws.cell(footer_t3, 2), round(dim_avgs.mean(), 2), pct=True, bold=True)

    autofit(ws)

    # Ancho fijo para columna A
    ws.column_dimensions["A"].width = 41

    wb.save(file_path)
    print(f"\nExcel report generated: {file_path}") 